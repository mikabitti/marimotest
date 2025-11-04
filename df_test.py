import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill


# df = pd.read_clipboard(sep="\t")
# print(df)

# data = {
#     'cust': ["John", "John", "John", "Marjanna", "Marjanna", "Marjanna"],
#     'prod': ['apple', 'apple', 'apple', 'banana', 'banana', 'banana'],
#     'amount': [1, 1, 2, 3, 2, 4],
#     'type': ['Vel', 'Vel', 'Hyv', 'Hyv', 'Vel', 'Vel']
# }
# df = pd.DataFrame(data)
df = pd.read_excel("test_data.xlsx")

# Create new column by subtracting type 'c' from type 'd' within same cust and prod groups


def calculate_adjusted_amount(group: pd.DataFrame):
    # Separate d and c types
    d_mask = group['type'] == 'd'
    c_mask = group['type'] == 'c'

    # Initialize result with zeros
    result = pd.Series(0, index=group.index, dtype=int)

    # Sum of c values to subtract
    c_total = group.loc[c_mask, 'amount'].sum()

    # Process d values - subtract c_total from them sequentially
    remaining_c = c_total

    for idx in group[d_mask].index:
        d_value = group.loc[idx, 'amount']
        subtraction = min(d_value, remaining_c)
        result.loc[idx] = max(0, d_value - subtraction)
        remaining_c -= subtraction

    # All c values become 0 (they're being subtracted from d values)
    # result already has 0 for c values from initialization

    return result


df['adjusted_amount'] = df.groupby(['cust', 'prod'], group_keys=False).apply(
    calculate_adjusted_amount)

print(df)

# Save to Excel with formatting
output_file = "df_test.xlsx"
df.to_excel(output_file, index=False)

# Load the workbook and apply conditional formatting
wb = openpyxl.load_workbook(output_file)
ws = wb.active

# Define color fills
red_fill = PatternFill(start_color="FF0000",
                       end_color="FF0000", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00",
                          end_color="FFFF00", fill_type="solid")
green_fill = PatternFill(start_color="00FF00",
                         end_color="00FF00", fill_type="solid")

# Find column indices (Excel is 1-indexed)
type_col = df.columns.get_loc('type') + 1
amount_col = df.columns.get_loc('amount') + 1
adjusted_col = df.columns.get_loc('adjusted_amount') + 1

# Apply formatting to each row
for idx, row in df.iterrows():
    if row['type'] == 'd':
        excel_row = idx + 2  # +1 for header, +1 for 0-indexed to 1-indexed

        # Determine color based on adjusted_amount
        if row['adjusted_amount'] == 0:
            # Red: adjusted amount is 0
            ws.cell(row=excel_row, column=type_col).fill = red_fill
        elif row['adjusted_amount'] == row['amount']:
            # Green: adjusted amount equals original amount
            ws.cell(row=excel_row, column=type_col).fill = green_fill
        else:
            # Yellow: adjusted amount is between 0 and original amount
            ws.cell(row=excel_row, column=type_col).fill = yellow_fill

wb.save(output_file)
print(f"\nExcel file saved with conditional formatting: {output_file}")
